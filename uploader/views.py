from django.shortcuts import render
from .models import Culturatemp
from openpyxl import load_workbook
from .forms import Culturatempform
from biao.forms import Culturaform
from biao.models import Cultura
from extracterpdf.models import Destinos
import os
import shutil
from website.settings import BASE_DIR



def index(request):
    if request.user.is_authenticated:
        uploader = request.user.groups.filter(name='Uploaders').exists()
        if uploader:
            return render(request, 'uploader/index.html', {'uploader': uploader,
                                                           'var1': 0})
        return render(request, 'biao/hrdb.html', {'uploader': uploader})
    return render(request, 'biao/hrdb_visitor.html')


def consolider(request, pk = None):
    if request.user.is_authenticated:
        uploader = request.user.groups.filter(name='Uploaders').exists()
        if uploader:
            primeira_cult = Culturatemp.objects.all().order_by('data', 'coleta').first()
            if primeira_cult is None:
                return render(request, 'uploader/erro.html',
                              {'erro': 'Sem novas culturas no banco de dados',
                               'uploader': uploader})
            if pk is not None:
                try:
                    cult = Culturatemp.objects.get(pk=pk)
                except:
                    return render(request, 'uploader/erro.html', {'erro': 'Cultura ' + str(pk) + ' não encontrada no banco de dados', 'uploader': uploader})
            else:
                cult = primeira_cult

            c = Cultura()
            for field in primeira_cult._meta.fields:
                if field.primary_key == True:
                    continue
                setattr(c, field.name, getattr(primeira_cult, field.name))

            if request.method == 'POST':
                form = Culturaform(request.POST, instance=c)
                if form.is_valid():
                    form.save()
                    primeira_cult.delete()
                    return render(request, 'biao/hrdb.html', {'uploader' : uploader})
            else:
                form = Culturaform(instance=c)

            return render(request, 'uploader/consolider2.html', {'uploader': uploader, 'primeira_cult': cult, 'form': form})

        return render(request, 'biao/hrdb.html')

    return render(request, 'biao/hrdb_visitor.html')


def confirmer(request, pk):
    uploader = request.user.groups.filter(name='Uploaders').exists()

    try:
        confirmavel = Culturatemp.objects.get(pk=pk)
        botao = None
        if request.method == 'GET':
            if request.GET.get('action') == 'del':
                botao = 'delete'
            if request.GET.get('action') == 'consolida':
                botao = 'consolida'
            if request.GET.get('action') == 'all':
                botao = 'consolidatudo'

        return render(request, 'uploader/confirmer.html', {'primeira_cult': confirmavel,
                                                           'uploader':uploader,
                                                           'botao': botao})
    except:
        return render(request, 'uploader/erro.html', {'erro': 'Não encontrada no banco de dados',
                                                      'uploader': uploader})


def confirmado(request, pk):
    uploader = request.user.groups.filter(name='Uploaders').exists()
    if request.GET.get('action') == 'del':
        primeira_cult = Culturatemp.objects.get(pk=pk)
        primeira_cult.delete()
        conclusao = 'apagado'
    elif request.GET.get('action') == 'all':
        conclusao = 'consolidadotudo'
        todas = Culturatemp.objects.all()
        for primeira_cult in todas:
            c = Cultura()
            for field in primeira_cult._meta.fields:
                if field.primary_key == True:
                    continue
                setattr(c, field.name, getattr(primeira_cult, field.name))
            primeira_cult.delete()
            c.save()
        primeira_cult = None
    elif request.GET.get('action') == 'consolida':
        primeira_cult = Culturatemp.objects.get(pk=pk)
        conclusao = 'consolidado'
        c = Cultura()
        for field in primeira_cult._meta.fields:
            if field.primary_key == True:
                continue
            setattr(c, field.name, getattr(primeira_cult, field.name))
        primeira_cult.delete()
        c.save()
    else:
        return render(request, 'uploader/erro.html',
                      {'erro': 'Erro na confirmação, reinicie o Consolider,',
                       'uploader': uploader})
    return render(request, 'uploader/confirmado.html', {'primeira_cult': primeira_cult,
                                                        'uploader': uploader,
                                                        'conclusao': conclusao})



def deleter(request):
    uploader = request.user.groups.filter(name='Uploaders').exists()
    pk = request.GET.get("p")
    noexist = False
    if pk is not None and pk != '':
        try:
            primeira_cult = Cultura.objects.get(pk=pk)
        except:
            primeira_cult = None
            noexist = True
    else:
        primeira_cult = None
    return render(request, 'uploader/deleter.html', {'primeira_cult': primeira_cult,
                                                     'uploader': uploader,
                                                     'noexist':noexist,
                                                     'pk':pk})


def deletado(request, pk):
    uploader = request.user.groups.filter(name='Uploaders').exists()
    try:
        condenada = Cultura.objects.get(pk=pk)
        nome = condenada.nome
        condenada.delete()
        confirma_delete = True
        noexist = False
    except:
        noexist = True
        confirma_delete = False
    return render(request, 'uploader/deleter.html', {'confirma_delete': confirma_delete,
                                                     'uploader': uploader,
                                                     'nome':nome,
                                                     'pk':pk,
                                                     'noexist':noexist})


def read(request):
    if request.user.is_authenticated:
        uploader = request.user.groups.filter(name='Uploaders').exists()
        if uploader:
            wb = load_workbook('c:/Users/gianr/PycharmProjects/django3/website/consolidadoDF1.xlsx')
            ws = wb['Sheet1']
            var1 = 0
            for row in ws.iter_rows(min_row=2):
                var1 += 1
                linha = Culturatemp(
                    filename=row[0].value,
                    nome=row[1].value,
                    prot=row[2].value,
                    medico=row[3].value,
                    data=row[4].value,
                    unid=row[5].value,
                    coleta=row[6].value,
                    material=row[7].value,
                    mat_especifico=row[8].value,
                    resultado=row[9].value,
                    tipo=row[10].value,
                    testes=row[11].value,
                    falha=row[12].value,
                    ami=row[13].value,
                    amp=row[14].value,
                    asb=row[15].value,
                    atm=row[16].value,
                    caz=row[17].value,
                    cip=row[18].value,
                    cli=row[19].value,
                    cpm=row[20].value,
                    cro=row[21].value,
                    ctn=row[22].value,
                    eri=row[23].value,
                    ert=row[24].value,
                    gen=row[25].value,
                    imi=row[26].value,
                    lin=row[27].value,
                    mer=row[28].value,
                    nit=row[28].value,
                    nor=row[30].value,
                    oxa=row[31].value,
                    pen=row[32].value,
                    pol=row[33].value,
                    ppt=row[34].value,
                    str=row[35].value,
                    sut=row[36].value,
                    tei=row[37].value,
                    tet=row[38].value,
                    van=row[39].value,
                )
                linha.save()
            uploader = request.user.groups.filter(name='Uploaders').exists()
            primeira_cult = Culturatemp.objects.all().order_by('data', 'coleta').first()
            if primeira_cult is None:
                return render(request, 'uploader/erro.html',
                              {'erro': 'XLSX sem novas entradas.',
                               'uploader': uploader})
            form = Culturatempform(request.POST or None, initial={"nome": primeira_cult.nome})
            return render(request, 'uploader/consolider2.html', {'var1': var1,
                                                                 'primeira_cult': primeira_cult,
                                                                 'uploader':uploader,
                                                                 'form':form,
                                                                 'deletado': False})
        return render(request, 'biao/hrdb.html', {'uploader':uploader})
    return render(request, 'biao/hrdb_visitor.html')




def consolidafiles(request):
    if request.user.is_authenticated:
        uploader = request.user.groups.filter(name='Uploaders').exists()
        if uploader:
            pre_path_pdfs = os.path.join(BASE_DIR, "static", "extracterpdf")
            path_pdfs = os.path.join(pre_path_pdfs, "pdfs_extraidos")
            #path_pdfs = r"extracterpdf\static\pdfs_extraidos"
            pre_path_pdfs_destino = os.path.join(BASE_DIR, "static", "biao")
            path_pdfs_destino = os.path.join(pre_path_pdfs_destino,  "pdfs")
            #path_pdfs_destino = r"biao\static\biao\pdfs"
            n = 0
            print(BASE_DIR)
            print(path_pdfs)
            print(path_pdfs_destino)
            for n, filename in enumerate(os.listdir(path_pdfs), 1):
                origem = os.path.join(path_pdfs, filename)
                destino = os.path.join(path_pdfs_destino, filename)
                shutil.move(origem, destino)

                condenados = Destinos.objects.all()
                for consolidado in condenados:
                    consolidado.delete()

            return render(request, 'uploader/consolidafiles.html', {'uploader': uploader, 'n': n})
        return render(request, 'biao/hrdb.html', {'uploader': uploader})
    return render(request, 'biao/hrdb_visitor.html')
