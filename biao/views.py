from .models import Cultura
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Count, Min, Max
from datetime import date
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect
from dateutil.relativedelta import relativedelta
from django.views.generic import View
from .forms import UserForm
from openpyxl import load_workbook
import os
from website.settings import BASE_DIR


def index(request):
    if not request.user.is_authenticated:
        return render(request, 'biao/hrdb_visitor.html')
    else:
        nome = request.GET.get("q") if request.GET.get("q") else ''
        resultado = request.GET.get("r") if request.GET.get("r") else ''
        material = request.GET.get("u") if request.GET.get("u") else ''
        mat_especifico = request.GET.get("v") if request.GET.get("v") else ''
        positivas = 'on' if request.GET.get("positivas", '') == 'on' else 'off'
        queryp = 'negativa' if request.GET.get("positivas", '') == 'on' else '@'
        ordem = ('-pk','-data', '-coleta') if request.GET.get("id_order", '') == 'on' else ('-data', '-coleta')
        id_order = request.GET.get("id_order", '')
        if not mat_especifico:
            if nome or resultado or material or positivas == 'on':
                todas_culturas = Cultura.objects.filter(
                    Q(nome__icontains=nome) & Q(resultado__icontains=resultado) & ~Q(resultado=queryp) & Q(
                        material__icontains=material)).order_by(*ordem)
            else:
                todas_culturas = Cultura.objects.all().order_by(*ordem)
        else:
            todas_culturas = Cultura.objects.filter(
                Q(nome__icontains=nome) & Q(resultado__icontains=resultado) & ~Q(resultado=queryp) & Q(
                    material__icontains=material) & Q(mat_especifico__icontains=mat_especifico)).order_by(*ordem)

        # paginator
        page = request.GET.get('page', 1)
        paginator = Paginator(todas_culturas, 30)
        try:
            culturas = paginator.page(page)
        except PageNotAnInteger:
            culturas = paginator.page(1)
        except EmptyPage:
            culturas = paginator.page(paginator.num_pages)
        uploader = request.user.groups.filter(name='Uploaders').exists()
        return render(request, 'biao/index.html',
                      {'todas_culturas': culturas,
                       'id_order': id_order,
                       'nome': nome,
                       'resultado': resultado,
                       'positivas': positivas,
                       'material': material,
                       'mat_especifico': mat_especifico,
                       'uploader':uploader})


def detail(request, pk):
    if not request.user.is_authenticated:
        return render(request, 'biao/hrdb_visitor.html')
    else:
        cultura = Cultura.objects.get(pk=pk)
        uploader = request.user.groups.filter(name='Uploaders').exists()
        return render(request, 'biao/detail.html', {'cultura': cultura,
                                                    'uploader':uploader})


def index_filtered(request, param):
    if not request.user.is_authenticated:
        return render(request, 'biao/hrdb_visitor.html')
    else:
        ordem = ('-data', '-coleta')
        id_order = ''
        todas_culturas = Cultura.objects.filter(nome=param).order_by(*ordem)

        # paginator
        page = request.GET.get('page', 1)
        paginator = Paginator(todas_culturas, 30)
        try:
            culturas = paginator.page(page)
        except PageNotAnInteger:
            culturas = paginator.page(1)
        except EmptyPage:
            culturas = paginator.page(paginator.num_pages)
        uploader = request.user.groups.filter(name='Uploaders').exists()
        return render(request, 'biao/index.html',
                      {'todas_culturas': culturas,
                       'nome': param,
                       'resultado': '',
                       'positivas': 'off',
                       'material': '',
                       'mat_especifico': '',
                       'uploader':uploader,
                       'id_order': id_order})


def hrdb(request):
    if request.user.is_authenticated:
        uploader = request.user.groups.filter(name='Uploaders').exists()
        return render(request, 'biao/hrdb.html', {'uploader' : uploader})
    else:
        return render(request, 'biao/hrdb_visitor.html')


def agente(request, agente):
    if not request.user.is_authenticated:
        return render(request, 'biao/hrdb_visitor.html')
    else:
        query_agente = Cultura.objects.filter(resultado=agente).order_by('-data')
        uploader = request.user.groups.filter(name='Uploaders').exists()
        return render(request, 'biao/agente.html', {'query_agente': query_agente,
                                                    'agente': agente,
                                                    'uploader':uploader})


def relacao(request, agente):
    if not request.user.is_authenticated:
        return render(request, 'biao/hrdb_visitor.html')
    else:
        dict_susceptibilidade = {}
        dict_total = {}

        query1 = Cultura.objects.exclude(
            Q(resultado='negativa') | Q(resultado=None) | Q(resultado='microbiota mista') | Q(
                resultado='microbiota normal') | Q(resultado='presença de coliformes') | Q(
                resultado='ausência de coliformes')
        ).values('resultado').annotate(num=Count('nome', distinct=True)).order_by('num').reverse()
        atbs = ['ami', 'amp', 'asb', 'atm', 'caz', 'cip', 'cli', 'cpm', 'cro', 'ctn', 'eri', 'ert', 'gen', 'imi', 'lin',
                'mer', 'nit', 'nor', 'oxa', 'pen', 'pol', 'ppt', 'str', 'sut', 'tei', 'tet', 'van']

        for atb in atbs:
            if agente != 'None':
                query2 = Cultura.objects.filter(
                    Q(resultado=agente) & ~Q(**{atb: None})
                ).values(atb).annotate(
                    num=Count('nome', distinct=True)).order_by('resultado')

            for item1 in query1:
                dict_total[item1['resultado']] = item1['num']

            if agente != 'None':
                r = 0
                s = 0
                for item in query2:
                    if item[atb] == 'r':
                        r = item['num']
                    elif item[atb] == 's':
                        s = item['num']
                testados = r + s
                tx = "{:4.1f}".format(100 * s / (r + s)) if testados != 0 else '-'
                dict_susceptibilidade[atb] = [tx, testados]
        uploader = request.user.groups.filter(name='Uploaders').exists()
        return render(request, 'biao/relacao.html', {'total': dict_total,
                                                     'susceptibilidade': dict_susceptibilidade,
                                                     'agente': agente,
                                                     'uploader': uploader})


def mdrs(request):
    if not request.user.is_authenticated:
        return render(request, 'biao/hrdb_visitor.html')
    else:
        limites = Cultura.objects.all().aggregate(Min('data'), Max('data'))

        query1 = Cultura.objects.exclude(
            Q(testes=None)).values('testes', 'data', 'nome').annotate(num=Count('nome', distinct=True))

        esbl = {}
        mrsa = {}
        kpc = {}
        data = limites['data__min']
        if data is not None and limites['data__max'] is not None:
            while data <= limites['data__max']:
                esbl[data.strftime("%b/%Y")] = 0
                mrsa[data.strftime("%b/%Y")] = 0
                kpc[data.strftime("%b/%Y")] = 0
                data += relativedelta(months=1)

        rep_esbl = []
        rep_mrsa = []
        rep_kpc = []
        for item in query1:
            if item['testes'] == 'esbl' and item['nome'] not in rep_esbl:
                rep_esbl.append(item['nome'])
                try:
                    esbl[item['data'].strftime("%b/%Y")] += 1
                except KeyError:
                    esbl[item['data'].strftime("%b/%Y")] = 1

            elif item['testes'] == 'mrsa' and item['nome'] not in rep_mrsa:
                rep_mrsa.append(item['nome'])
                try:
                    mrsa[item['data'].strftime("%b/%Y")] += 1
                except KeyError:
                    mrsa[item['data'].strftime("%b/%Y")] = 1
            elif item['testes'] == 'kpc' and item['nome'] not in rep_kpc:
                rep_kpc.append(item['nome'])
                try:
                    kpc[item['data'].strftime("%b/%Y")] += 1
                except KeyError:
                    kpc[item['data'].strftime("%b/%Y")] = 1

        esbl_l = [(k, v) for k, v in esbl.items()]
        esbl_l.reverse()
        mrsa_l = [(k, v) for k, v in mrsa.items()]
        mrsa_l.reverse()
        kpc_l = [(k, v) for k, v in kpc.items()]
        kpc_l.reverse()
        uploader = request.user.groups.filter(name='Uploaders').exists()
        return render(request, 'biao/mdrs.html', {'esbl': esbl_l,
                                                  'mrsa': mrsa_l,
                                                  'kpc': kpc_l,
                                                  'uploader': uploader})


def login_user(request):
    if request.user.is_authenticated:
        uploader = request.user.groups.filter(name='Uploaders').exists()
        return render(request, 'biao/hrdb.html', {'uploader': uploader})
    else:
        if request.method == "POST":
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(username=username, password=password)
            if user is not None:
                if user.is_active:
                    login(request, user)
                    uploader = request.user.groups.filter(name='Uploaders').exists()
                    return render(request, 'biao/hrdb.html', {'uploader':uploader})
                else:
                    return render(request, 'biao/login.html', {'error_message': 'Your account has been disabled'})
            else:
                return render(request, 'biao/login.html', {'error_message': 'Invalid login'})
        return render(request, 'biao/login.html')


def logout_user(request):
    if not request.user.is_authenticated:
        return render(request, 'biao/hrdb_visitor.html')
    else:
        logout(request)
        form = UserForm(request.POST or None)
        context = {
            "form": form,
        }
        return render(request, 'biao/hrdb_visitor.html', context)


def register(request):
    if request.user.is_authenticated:
        uploader = request.user.groups.filter(name='Uploaders').exists()
        return render(request, 'biao/hrdb.html', {'uploader':uploader})
    else:
        form = UserForm(request.POST or None)
        if form.is_valid():
            user = form.save(commit=False)
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user.set_password(password)
            user.save()
            user = authenticate(username=username, password=password)
            if user is not None:
                if user.is_active:
                    login(request, user)
                    uploader = request.user.groups.filter(name='Uploaders').exists()
                    return render(request, 'biao/hrdb.html', {'uploader':uploader})
        context = {
            "form": form,
        }
        return render(request, 'biao/registration_form.html', context)


def read(request):
    wb = load_workbook('c:/Users/gianr/PycharmProjects/django3/website/CULTURAS DEFINITIVO.xlsx')
    ws = wb['Sheet1']
    var1 = 0
    for row in ws.iter_rows(min_row=2):
        var1 += 1
        linha = Cultura(
            filename=row[0].value,
            nome=row[1].value,
            prot=row[2].value,
            medico=row[3].value,
            data=row[4].value,
            unid=row[5].value,
            coleta=row[6].value,
            material=row[7].value,
            mat_especifico=row[8].value,
            resultado=row[9].value,
            tipo=row[10].value,
            testes=row[11].value,
            falha=row[12].value,
            ami=row[13].value,
            amp=row[14].value,
            asb=row[15].value,
            atm=row[16].value,
            caz=row[17].value,
            cip=row[18].value,
            cli=row[19].value,
            cpm=row[20].value,
            cro=row[21].value,
            ctn=row[22].value,
            eri=row[23].value,
            ert=row[24].value,
            gen=row[25].value,
            imi=row[26].value,
            lin=row[27].value,
            mer=row[28].value,
            nit=row[28].value,
            nor=row[30].value,
            oxa=row[31].value,
            pen=row[32].value,
            pol=row[33].value,
            ppt=row[34].value,
            str=row[35].value,
            sut=row[36].value,
            tei=row[37].value,
            tet=row[38].value,
            van=row[39].value,
        )
        linha.save()
    uploader = request.user.groups.filter(name='Uploaders').exists()
    return render(request, 'biao/read.html', {'var1': var1,
                                              'uploader':uploader})


def limpahopsital(request):
    if request.user.is_authenticated:
        uploader = request.user.groups.filter(name='Uploaders').exists()
        if uploader:
            print(request.method)
            if request.method == 'GET':
                print(request.GET.get('decisao'))
                if request.GET.get('decisao') == 'limpa':
                    Cultura.objects.all().delete()
                    pre_path_pdfs_destino = os.path.join(BASE_DIR, "static", "biao")
                    path_pdfs_destino = os.path.join(pre_path_pdfs_destino, "pdfs")
                    for filename in os.listdir(path_pdfs_destino):
                        os.remove(os.path.join(path_pdfs_destino, filename))
                    print('APAGOU TUDO')
                    return render(request, 'biao/limpado.html', {'uploader': uploader})
                elif request.GET.get('decisao') == 'desiste':
                    return render(request, 'biao/hrdb.html', {'uploader': uploader} )
                else:
                    return render(request, 'biao/limpahospital.html', {'uploader': uploader})
            else:
                return render(request, 'biao/limpahospital.html', {'uploader': uploader})
    return render(request, 'biao/hrdb.html')
